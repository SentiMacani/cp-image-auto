import os
import time
from datetime import datetime

import requests
from dotenv import load_dotenv
from eth_account import Account
from fake_useragent import UserAgent
from faker import Faker
from mnemonic import Mnemonic
from openpyxl import Workbook, load_workbook
from web3 import Web3

from utils.constants import PROVIDER_URL, EXCEL_FILE_PATH


class Register:
    def __init__(self, stop_flag):
        load_dotenv()
        self.ua = UserAgent()
        self.faker = Faker()
        self.provider_url = PROVIDER_URL
        self.file_path = EXCEL_FILE_PATH
        self.stop_flag = stop_flag

    def write_to_excel(self, headers, data):
        max_retries = 3
        for attempt in range(max_retries):
            if self.stop_flag.is_set():
                print("Stopping Excel write operation due to interrupt.")
                return
            try:
                file_exists = os.path.isfile(self.file_path)
                wb = load_workbook(self.file_path) if file_exists else Workbook()
                ws = wb.active
                if not file_exists:
                    ws.append(headers)
                ws.append(data)
                wb.save(self.file_path)
                break
            except PermissionError:
                if attempt < max_retries - 1:
                    time.sleep(1)
                else:
                    print(f"Failed to write to Excel after {max_retries} attempts.")
            except Exception as e:
                print(f"Error writing to Excel: {e}")
                break

    def get_account(self):
        mnemo = Mnemonic("english")
        mnemonic = mnemo.generate(strength=128)
        web3 = Web3(Web3.HTTPProvider(self.provider_url))
        web3.eth.account.enable_unaudited_hdwallet_features()
        try:
            account = Account.from_mnemonic(mnemonic)
            return {
                "address": account.address,
                "mnemonics": mnemonic,
                "privateKey": account._private_key.hex(),
            }
        except Exception as e:
            print(f"Account generation error: {e}")
            return None

    def make_request(self, method, url, headers, json=None, proxies=None):
        if self.stop_flag.is_set():
            print("Stopping request due to interrupt.")
            return None
        try:
            with requests.Session() as session:
                response = session.request(method, url, headers=headers, json=json, proxies=proxies, timeout=10)
            response.raise_for_status()
            return response.json()
        except requests.exceptions.RequestException as e:
            print(f"Request error: {e}")
            return None

    def login(self, account, user_agent, proxies):
        headers = {
            "User-Agent": user_agent,
            "accept": "application/json, text/plain, */*",
            "content-type": "text/plain",
            "Referer": "https://www.creoplay.app/",
        }
        url = "https://m84784bbba.execute-api.ap-southeast-1.amazonaws.com/production/authentication/wallet/login"
        payload = {"wallet_id": account}
        data = self.make_request("POST", url, headers, json=payload, proxies=proxies)
        return data["data"]["api_token"] if data else None

    def register_username(self, api_token, user_agent, proxies):
        headers = {
            "User-Agent": user_agent,
            "accept": "application/json, text/plain, */*",
            "content-type": "text/plain",
            "Authorization": f"Bearer {api_token}",
            "Referer": "https://www.creoplay.app/",
        }
        url = "https://m84784bbba.execute-api.ap-southeast-1.amazonaws.com/production/authentication/username"
        username = f"{self.faker.first_name()}_{self.faker.last_name()}"
        payload = {"username": username}
        self.make_request("PUT", url, headers, json=payload, proxies=proxies)
        return username

    def register_password(self, api_token, user_agent, proxies):
        headers = {
            "User-Agent": user_agent,
            "accept": "application/json, text/plain, */*",
            "content-type": "text/plain",
            "Authorization": f"Bearer {api_token}",
            "Referer": "https://www.creoplay.app/",
        }
        url = "https://m84784bbba.execute-api.ap-southeast-1.amazonaws.com/production/authentication/password"
        payload = {"password": "Q!w2e3r4t5y6", "password_confirmation": "Q!w2e3r4t5y6"}
        self.make_request("PUT", url, headers, json=payload, proxies=proxies)
        return "Q!w2e3r4t5y6"

    def get_login_key(self, api_token, user_agent, proxies):
        headers = {
            "User-Agent": user_agent,
            "accept": "application/json, text/plain, */*",
            "content-type": "text/plain",
            "Authorization": f"Bearer {api_token}",
            "Referer": "https://www.creoplay.app/",
        }
        url = "https://m84784bbba.execute-api.ap-southeast-1.amazonaws.com/production/authentication/login-key"
        data = self.make_request("POST", url, headers, proxies=proxies)
        return data["data"]["login_key"] if data else None

    def get_data(self, api_token, user_agent, proxies):
        headers = {
            "User-Agent": user_agent,
            "accept": "application/json, text/plain, */*",
            "content-type": "text/plain",
            "authorization": f"Bearer {api_token}",
            "Referer": "https://www.creoplay.app/",
        }
        url = "https://m84784bbba.execute-api.ap-southeast-1.amazonaws.com/production/authentication"
        return self.make_request("GET", url, headers, proxies=proxies)

    def register_user(self, proxy_url, index):
        print(f"Register User Thread {index} started")
        try:
            user_agent = self.ua.random
            account = self.get_account()

            # Add frequent checks for threading.current_thread().stop_requested
            # For example:
            if self.stop_flag.is_set():
                print(f"Thread {index} stopping due to interrupt.")
                return

            proxies = {
                "http": f"http://{proxy_url}",
                "https": f"http://{proxy_url}",
            }

            api_token = self.login(account["address"], user_agent, proxies)
            if not api_token or self.stop_flag.is_set():
                return

            username = self.register_username(api_token, user_agent, proxies)
            password = self.register_password(api_token, user_agent, proxies)
            login_key = self.get_login_key(api_token, user_agent, proxies)
            user_data = self.get_data(api_token, user_agent, proxies)
            user_reg_data = datetime.now().strftime("%Y-%m-%d")

            if not all([username, password, login_key, user_data]) or self.stop_flag.is_set():
                return

            headers = [
                "username", "password", "address", "login_key", "mnemonics", "privateKey",
                "creo_id", "hash_id", "login_key_expired_at", "is_account_completed",
                "proxy_url", "user_agent", "creation_date"
            ]
            registration_data = [
                f"{username}#{user_data['data']['hash_id']}", password, account["address"],
                login_key, account["mnemonics"], account["privateKey"],
                user_data["data"]["creo_id"], user_data["data"]["hash_id"],
                user_data["data"]["login_key_expired_at"], user_data["data"]["is_account_completed"],
                proxy_url, user_agent, user_reg_data
            ]
            self.write_to_excel(headers, registration_data)

            print(f'Register User Thread {index} finished')
        except Exception as e:
            print(f"Error in Register User Thread {index}: {e}")
        finally:
            if self.stop_flag.is_set():
                print(f"Register User Thread {index} stopped due to interrupt")
