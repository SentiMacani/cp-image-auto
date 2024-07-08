import unittest
import subprocess
import pyfiglet
import time
import threading
import random
import time
from web3 import Web3
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor, as_completed
from adbutils import adb
from collections import deque
from termcolor import colored
from appium import webdriver
from appium.options.android import UiAutomator2Options
from appium.options.common import AppiumOptions
from game_loop_3 import Evermore_Nights_Loop_3
from utils.getLoginKey import *
from utils.execp_handler import *
from utils.ip_loc import *
from utils.readFromCsv import *
from utils.android_actions import *
from utils.app_actions import *
from utils.constants import *
from utils.mobile_capabilities import *
from utils.tools import get_random_wallet,get_all_wallets,prepare_transaction,send_transaction
from modules.registration import Register
#from game_loop_1 import Evermore_Nights_Loop_1
# from loop_5 import Evermore_Nights_Loop_5
from game_loop_1_y8 import Evermore_Nights_Loop_1
from eth_account import Account
from utils.networks import get_network_by_name

network=get_network_by_name(ENV)
provider_url=network['rpc']
web3 = Web3(Web3.HTTPProvider(provider_url))


nft_contract = web3.eth.contract(address=NFT_CONTRACTS[ENV], abi=NFT_CONTRACT_ABI)
nft_marketplace_contract =web3.eth.contract(address=MARKETPLACE_CONTRACTS[ENV], abi=MARKETPLACE_ABI)
quickswap_contract=web3.eth.contract(address=QUICKSWAP_CONTRACTS[ENV],abi=QUICKSWAP_ROUTER_ABI)



## ADB and Appium Handlers

def setup_driver(device, port):
    capabilities = generalized_phone.copy()
    capabilities["deviceName"] = str(device.serial)
    capabilities["udid"] = str(device.serial)
    appium_server_url_local = appium_server_url+":"+str(port)
    driver = webdriver.Remote(command_executor=appium_server_url_local, options=AppiumOptions().load_capabilities(capabilities))
    active_drivers.append(driver)
    return driver

def teardown_driver(driver):
    if driver:
        driver.quit()

def get_device_resolution(device):
    try:
        output = subprocess.check_output(['adb', '-s', device.serial, 'shell', 'wm', 'size']).decode('utf-8').strip()
        size = output.split(':')[1].strip()
        width, height = map(int, size.split('x'))
        return width, height
    except Exception as e:
        print(f"Error getting screen size via ADB for device {device}: {e}")
        return None, None
    
def get_eth_balance(address):
    try:
        balance_wei = web3.eth.get_balance(address)
        return balance_wei
    except Exception as e:
        print(f"Error getting balance: {e}")
        return None
    
    
def send_eth(sender_private_key, sender_public_key, to_address, amount_in_wei):
    try:
        nonce = web3.eth.get_transaction_count(sender_public_key)
        while True:
            try:
                transaction = {
                    'to': to_address,
                    'value': int(amount_in_wei),
                    'gas': 21000,  
                    'gasPrice': web3.to_wei('20', 'gwei'), 
                    'nonce': nonce,
                    'chainId': web3.eth.chain_id
                }
                signed_txn = web3.eth.account.sign_transaction(transaction, sender_private_key)
                txn_hash = web3.eth.send_raw_transaction(signed_txn.rawTransaction)
                txn_receipt = web3.eth.wait_for_transaction_receipt(txn_hash)
                txn_url = f"{EXPLORER_URL}tx/{txn_hash.hex()}"
                print(f"Transaction successful: {txn_url}")
                return txn_receipt

            except ValueError as ve:
                error_message = ve.args[0].get('message', '')
                if 'nonce too low' in error_message:
                    print(f"Nonce too low, fetching the latest nonce and trying again...")
                    nonce = web3.eth.get_transaction_count(sender_public_key)
                else:
                    raise ve
    except Exception as e:
        print(f"An error occurred: {e}")
        return None
#4a5c062b368b2f679b09697934514ec9f54a9eb9a593af21a43ddeb02311f720
#0x08a74A0075a2C3A786A84439812a141C6C8b73f3 
pri='4a5c062b368b2f679b09697934514ec9f54a9eb9a593af21a43ddeb02311f720'
pub='0x08a74A0075a2C3A786A84439812a141C6C8b73f3'
        
def transfer_okb(private_key,public_key):
    to_private_key, to_public_key = get_random_wallet()           
    while to_public_key == public_key:
        to_private_key, to_public_key = get_random_wallet()
    from_balance = get_eth_balance(public_key)
    max_amount_rotation = web3.to_wei(MAX_AMOUNT_ROTATION, 'ether')
    max_amount = from_balance if from_balance < max_amount_rotation else max_amount_rotation
    amount = round(random.uniform(0, float(max_amount)), 6)
    amount_in_wei = web3.from_wei(amount, 'ether')
    print(f"Sending {amount_in_wei} from {public_key} to {to_public_key}")
    tx_params = prepare_transaction(public_key,int(amount))
    tx_params["to"] = to_public_key
    return send_transaction(tx_params,private_key)
            
def swap_quickswap_okx_usdc(private_key,public_key):
    amount = round(random.uniform(0.001, 0.003), 4)
    amount_in_wei = int(amount * 10**18)
    print(f"Swapping OKX -> USDC: {amount} ")
    params = {
        'tokenIn': TOKENS_PER_CHAIN[ENV]["WBNB"],  # WETH
        'tokenOut': TOKENS_PER_CHAIN[ENV]["USDC"],  # USDC
        'fee': 3000,  # Default fee tier (0.3%)
        'recipient': public_key,  # Recipient address
        'deadline': 9999999999,  # Deadline (current time + 5 minutes)
        'amountIn': amount_in_wei,  # Amount In ETH
        'amountOutMinimum': 10,  # Amount Out Minimum (set to 0 for this example, but should be set to avoid slippage)
        'sqrtPriceLimitX96': 0  # SqrtPriceLimitX96 (set to 0 for no price limit)
    }

    tx_params =prepare_transaction(public_key,value=amount_in_wei)
    transaction = quickswap_contract.functions.exactInputSingle(params).build_transaction(tx_params)    
    return send_transaction(transaction,private_key)

def mint_nft(private_key,public_key):
        print(f"Mint NFT on NFT CONTRACT.")
        transaction = nft_contract.functions.mintToken(
            "some_data"
        ).build_transaction(prepare_transaction(public_key))
        return send_transaction(transaction,private_key)
    
def create_market_sale(private_key,public_key):
    print(f"Buy NFT on MarketPlace.")

    all_nfts = nft_marketplace_contract.functions.fetchMarketToken().call()
    if len(all_nfts) > 0:
        random_id = random.randint(0, len(all_nfts) - 1)
        item_id = all_nfts[random_id][0]  # Item Id
        price = all_nfts[random_id][5]  # Price
        tx_params = prepare_transaction(public_key,value=price)
        transaction = nft_marketplace_contract.functions.createMarketSale(
            nft_contract.address, item_id
        ).build_transaction(tx_params)
        send_transaction(transaction,private_key)
    else:
        print("No Nfts available for Sale")

def make_market_item(private_key,public_key):
    user_nfts = nft_contract.functions.balanceOf(public_key).call()
    if user_nfts > 0:
        print("Listing NFT on MarketPlace.")
        listingPrice = nft_marketplace_contract.functions.getListingPrice().call()
        random_nft_index = random.randint(0, user_nfts - 1)
        token_id = nft_contract.functions.tokenOfOwnerByIndex(public_key, random_nft_index).call()
        price = random.randint(1, 10)
        tx_params = prepare_transaction(public_key,value=listingPrice)
        transaction = nft_marketplace_contract.functions.MintMarketItem(nft_contract.address, token_id, price).build_transaction(tx_params)
        return send_transaction(transaction, private_key)
    else:
        return "No Nfts available for Sale"


function_map = {
    'transfer_okb': transfer_okb,
    'swap_quickswap_okx_usdc': swap_quickswap_okx_usdc,
    'mint_nft': mint_nft,
    'create_market_sale': create_market_sale,
    'make_market_item': make_market_item
}  

## Core Menu Controls 2 and 3 - @Shivansh TODO
def rotate_and_mix(num_rotations):
    print(f"Rotating and Mixing {num_rotations} to onchain wallets...")

    for i in range(num_rotations):
        try:
            from_private_key, from_public_key = get_random_wallet()
            random_function_name = random.choice(CLASSIC_ROUTES_MODULES_USING)
            function_map[random_function_name](from_private_key, from_public_key)
        except Exception as e:
            print(f"An error occurred during route {random_function_name} {i + 1}: {e}")
        # sleep_duration = random.randint(2, 7)
        # print(f"Sleeping for {sleep_duration} seconds before next transfer...")
        # time.sleep(sleep_duration)
             
def fund_wallets(private_key,public_key,min_amount):
    print(f"Funding {min_amount}:amounts to required wallets ")
    all_wallets=get_all_wallets()
    min_amount_wei=web3.from_wei(min_amount, 'ether')
    for i in range(len(all_wallets)):
        try:  
            to_public_key=all_wallets[i]
            print('to_public_key',to_public_key)
            to_balance = get_eth_balance(to_public_key)
            if to_balance<min_amount_wei:
                print(f"Sending {min_amount} from {public_key} to {to_public_key}")
                tx_hash=send_eth(private_key,public_key,to_public_key,min_amount_wei)
            else:
                print(f"{to_public_key} has sufficient balance ... Skipping")     
        except Exception as e:
            print(f"An error occurred during transfer {i + 1}: {e}")
             
            
        
def activate_historical_txns():
    print("Activating historical txns on onchain wallets...")

## Core Menu Controls 1 - Create CP Accounts

def generate_accounts(num_accounts):
    print(f"Generating {num_accounts} CreoPlay accounts and onchain Wallets...")
    register = Register(stop_flag)
    with open("data/proxy_links.txt", "r") as file:
        proxy_links = [line.strip() for line in file.readlines()]
    random.shuffle(proxy_links)
    proxy_queue = deque(proxy_links)
    accounts_per_thread = num_accounts // NUM_THREADS
    remainder = num_accounts % NUM_THREADS
    threads = []
    for i in range(NUM_THREADS):
        thread_account_count = accounts_per_thread + (1 if i < remainder else 0)
        thread_proxies = [proxy_queue.popleft() for _ in range(len(proxy_links) // NUM_THREADS)]
        if i < len(proxy_links) % NUM_THREADS:
            thread_proxies.append(proxy_queue.popleft())
        
        thread = threading.Thread(target=automate_registration_loop, 
                                  args=(register, thread_proxies, i, thread_account_count))
        threads.append(thread)
        thread.start()
    try:
        while any(thread.is_alive() for thread in threads):
            for thread in threads:
                thread.join(timeout=0.5)
            if stop_flag.is_set():
                break
    except KeyboardInterrupt:
        print("\nKeyboard interrupt received in main thread.")
        stop_flag.set()
    # Ensure all threads have stopped
    for thread in threads:
        thread.join()
    print("Account generation process completed or interrupted.") 


def automate_registration_loop(register, proxies, thread_index, account_count):
    accounts_created = 0
    while accounts_created < account_count and not stop_flag.is_set():
        try:
            proxy_url = random.choice(proxies)
            sleep_duration = random.randint(2, 10)
            print(f"Thread-{thread_index} using proxy traced to location {trace_proxy_connection(proxy_url)}, waiting {sleep_duration} seconds before next registration...")
            for _ in range(sleep_duration):
                if stop_flag.is_set():
                    print(f"Thread-{thread_index} stopping due to interrupt.")
                    return
                time.sleep(1)
            register.register_user(proxy_url, thread_index)
            accounts_created += 1
            print(f"Thread-{thread_index} created account {accounts_created}/{account_count}")
        except Exception as e:
            print(f"Error in Thread-{thread_index}: {e}")
            if stop_flag.is_set():
                return

## Core Menu Controls 4 - Ingame Automation

def load_accounts_from_excel(file_path):
    workbook = load_workbook(filename=file_path, read_only=True)
    sheet = workbook.active
    headers = [cell.value for cell in sheet[1]]
    accounts = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        account = {headers[i]: value for i, value in enumerate(row)}
        accounts.append(account)
    return accounts

def run_tests_for_device(device, accounts, port):
    width, height = get_device_resolution(device)
    for account in accounts:
        if stop_flag.is_set():
            print(f"Stopping tests for device {device.serial}")
            break
        username = account['username']
        mnemonic = account['mnemonics']
        address = account['address']
        if username and mnemonic and address:
            login_key = createLoginKey(mnemonic)
            print(f"Device {device.serial} (Port {port}) - Username: {username}")
            print(f"Device {device.serial} (Port {port}) - PubKey Addr: {address}")
            print(f"Device {device.serial} (Port {port}) - Login key: {login_key}")
            print(f"Device {device.serial} (Port {port}) - Screen Size: {width}x{height}")
            driver = setup_driver(device, port)
            try:
                # Run test_loop_1
                test_case_1 = Evermore_Nights_Loop_1('test_loop_1', driver=driver, login_key=login_key, username=username, pub_key=address, screen_width=width, screen_height=height, stop_flag=stop_flag)
                result_1 = unittest.TextTestRunner().run(unittest.TestSuite([test_case_1]))
                
                if result_1.wasSuccessful() and not stop_flag.is_set():
                    # Only run test_loop_3 if test_loop_1 was successful and stop_flag is not set
                    test_case_3 = Evermore_Nights_Loop_3('test_loop_3', driver=driver, login_key=login_key, username=username, pub_key=address, screen_width=width, screen_height=height, stop_flag=stop_flag)
                    unittest.TextTestRunner().run(unittest.TestSuite([test_case_3]))
                else:
                    print(f"Skipping test_loop_3 for device {device.serial} due to test_loop_1 failure or stop flag")
                
            except Exception as e:
                print(f"Caught exception and restarting with next account {e}")
            finally:
                teardown_driver(driver)
    

def activate_automation_loop():
    print("Activating In-game automation loop...")
    devices = adb.device_list()
    if not devices:
        print("No devices connected. Please connect devices and try again.")
        return
    file_path = r'data/registration_data.xlsx'  # Change this to your Excel file path
    accounts = load_accounts_from_excel(file_path)
    # Split accounts across devices
    accounts_per_device = [accounts[i::len(devices)] for i in range(len(devices))]
    threads = []
    for i, (device, device_accounts) in enumerate(zip(devices, accounts_per_device)):
        port = base_appium_port + i
        thread = threading.Thread(target=run_tests_for_device, args=(device, device_accounts, port))
        threads.append(thread)
        thread.start()
        
        # Add a random delay between 3 to 4 seconds before starting the next thread
        delay = random.uniform(3, 4)
        print(f"Waiting {delay:.2f} seconds before starting the next device...")
        time.sleep(delay)
    try:
        # Wait for all threads to complete
        for thread in threads:
            while thread.is_alive():
                thread.join(timeout=0.5)
                if stop_flag.is_set():
                    break
            if stop_flag.is_set():
                break
    except KeyboardInterrupt:
        print("\nKeyboard interrupt received. Stopping all threads...")
        stop_flag.set()
    # Ensure all threads have stopped
    for thread in threads:
        thread.join()
    print("Automation loop completed or interrupted.")


def print_menu():
    # Print ASCII art title
    title = pyfiglet.figlet_format("Genleap", font="slant")
    colored_title = colored(title, "cyan", attrs=["bold"])
    print(colored_title)

    # Print subheading
    subheading = "Customized for Creoplay"
    colored_subheading = colored(subheading, "yellow", attrs=["bold"])
    print(colored_subheading)
    print("\n")

    # Print menu options
    print("1. Generate CreoPlay accounts and onchain Wallets")
    print("2. Fund Wallets")
    print("3. Rotate and Mix transfers to onchain wallets")
    print("4. Activate historical txns on onchain wallets")
    print("5. Activate In-game automation loops")
    print("6. Exit")

def main():
    while not stop_flag.is_set():
        try:
            print_menu()
            choice = input("Enter your choice (1-5): ")

            if choice == '1':
                while True:
                    try:
                        num_accounts = int(input("How many accounts do you want to create? "))
                        if num_accounts > 0:
                            break
                        else:
                            print("Please enter a positive number.")
                    except ValueError:
                        print("Please enter a valid number.")
                generate_accounts(num_accounts)
            elif choice == '2':
                while True:
                    try:
                        private_key = input("Enter private Key")
                        public_key =(Account.from_key(private_key)).address
                        if public_key:
                            break
                    except ValueError:
                        print("Please enter a valid privateKey.")          
                while True:
                    try:
                        min_amount = float(input("Enter Amount for Funding "))
                        if min_amount > 0:
                            break
                        else:
                            print("Please enter a positive number.")
                    except ValueError:
                        print("Please enter a valid number.")
                fund_wallets(private_key,public_key,min_amount)
            elif choice == '3':
                while True:
                    try:
                        num_rotations = int(input("How many number of rotations ? "))
                        if num_rotations > 0:
                            break
                        else:
                            print("Please enter a positive number.")
                    except ValueError:
                        print("Please enter a valid number.")          
                rotate_and_mix(num_rotations)
            elif choice == '4':
                activate_historical_txns()
            elif choice == '5':
                activate_automation_loop()
            elif choice == '6':
                print("Exiting...")
                break
            else:
                print("Invalid choice. Please try again.")

            if not stop_flag.is_set():
                input("\nPress Enter to continue...")
        except KeyboardInterrupt:
            print("\nCtrl+C detected in main loop. Initiating graceful shutdown...")
            stop_flag.set()
        except Exception as e:
            print(f"An error occurred: {e}")
            stop_flag.set()

    print("Exiting program.")

if __name__ == "__main__":
    main()